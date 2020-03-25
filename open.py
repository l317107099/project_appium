import openpyxl
from openpyxl import Workbook

# wb = Workbook('公司名称.xlsx')
wb = openpyxl.load_workbook("home.xlsx")
# ws = wb['data']
ws = wb.active
a=0
for i in ws.rows:
    a+=1
    for j in i:
        print(j.value)
#

# print('{}行 {}列'.format(ws.max_row, ws.max_column))

